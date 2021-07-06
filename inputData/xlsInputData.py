from openpyxl import load_workbook

# from numpy import array
# from configuration import INPUT_FILE
from model.job import Job
from model.task import Task


class XslInputData:
    def __init__(self, input_file):
        file_excel = load_workbook(input_file, data_only=True)
        # print(file_excel.sheetnames)

        self._sheet1 = file_excel['Sheet1']
        self._sheet_task = file_excel['Task']
        self._sheet_matrix = file_excel['MatriceJ-T']
        self._sheet_job = file_excel['Job']
        self._jobs_num = int(self._sheet1['B1'].value)  # vado a leggere il numero di jobs
        self._tasks_num = int(self._sheet1['B2'].value)  # vado a leggere il numero di tasks
        self._capacity_batch = int(self._sheet1[
                                       'B3'].value)  # vado a leggere il numero di provette che possono essere processate contemporaneamente

        self._duration_t = []
        for i in range(self._tasks_num):
            # duration_t = sheet_task.cell(i + 2, 2).value
            self._duration_t.append(self._sheet_task.cell(i + 2, 2).value)

    def read_jobs(self):
        # Vado a leggere la matrice Job_x_Task
        matrix_TJ = [[0] * self._tasks_num for _ in range(self._jobs_num)]
        for i in range(self._jobs_num):
            for j in range(self._tasks_num):
                matrix_TJ[i][j] = self._sheet_matrix.cell(i + 2, j + 2).value
        # print(matrix_TJ)

        # Vado a leggere le durate dei task e le metto in un array
        # duration_t = array('I', range(tasks_num))

        # print(duration_t)

        jobs = []
        # Vado a crearmi la lista dei jobs.
        for i in range(self._jobs_num):
            list_task_job = []
            for j in range(self._tasks_num):
                # MODIFICA PER LE DURATE = 0
                if matrix_TJ[i][j] == 1:
                    list_task_job.append(
                        # Task(j, matrix_TJ[i][j] * duration_t[j]))   per ciascun jobs, vado a leggere e tenere in
                        Task(j, self._duration_t[j]))
                # memoria la durata dei task (la durata sarà 0 per i tasks non presenti in quello specifico jobs
            jobs.append(Job(i, self._sheet_job.cell(i + 2, 2).value, self._sheet_job.cell(i + 2, 3).value,
                            list_task_job))  # inizializzo il
            # jobs con il suo id, releaseTime, dueDate, e la lista dei suoi task e lo aggiungo alla lista dei jobs
        return jobs

    def read_capacity_batch(self):
        return self._capacity_batch

    def read_duration_task(self):
        return self._duration_t
