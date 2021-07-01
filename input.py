from openpyxl import load_workbook
from numpy import array
from configuration import INPUT_FILE
from task import *
from job import *

file_excel = load_workbook(INPUT_FILE, data_only=True)
print(file_excel.sheetnames)
# dal primo foglio della mia cartella excel
sheet1 = file_excel['Sheet1']
n = sheet1['B1'].value  # vado a leggere il numero di jobs
k = sheet1['B2'].value  # vado a leggere il numero di tasks
M = sheet1['B3'].value  # vado a leggere il numero di provette che possono essere processate contemporaneamente

# Vado a leggere la matrice Job_x_Task
sheet_matrix = file_excel['MatriceJ-T']
matrix_TJ = [[0] * k for _ in range(n)]

for i in range(n):
    for j in range(k):
        matrix_TJ[i][j] = sheet_matrix.cell(i + 2, j + 2).value

print(matrix_TJ)

# Vado a leggere le durate dei task e le metto in un array

sheet_task = file_excel['Task']
duration_t = array('I', range(k))

for i in range(k):
    duration_t[i] = sheet_task.cell(i + 2, 2).value

print(duration_t)

# Vado a crearmi la lista dei job.


sheet_job = file_excel['Job']
job = []

for i in range(n):
    list_task_job = []
    for j in range(k):
        list_task_job.append(Task(j, matrix_TJ[i][j] * duration_t[j]))  # per ciascun job, vado a leggere e tenere in
        # memoria la durata dei task (la durata sarà 0 per i tasks non presenti in quello specifico job
    job.append(Job(i, sheet_job.cell(i + 2, 2).value, sheet_job.cell(i + 2, 3).value, list_task_job))  # inizializzo il
    # job con il suo id, releaseTime, dueDate, e la lista dei suoi task e lo aggiungo alla lista dei job
